import io
from dateutil import parser
from PIL import Image as PILImage
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session
import pyrebase
import re
import firebase_admin
from firebase_admin import credentials, firestore, auth, storage
from datetime import datetime
import os
import shutil
import os.path
import openpyxl
from google.cloud import vision
from google.cloud import vision_v1
from openpyxl.styles import Font, Alignment
import requests
from requests.exceptions import ConnectionError

# Initialize Flask app
app = Flask(__name__)
app.secret_key = 'your_secret_key'
uid = ""
site_code=""
idToken = ""
globalCircleName = ""
employeeName = ""
azimuth =""

# Initialize Firebase
firebaseConfig = {
    "apiKey": "AIzaSyBiZWWI4vdWhNjwK1547w6dRHKBVZ9it24",
    "authDomain": "telecom-tower-performance-1.firebaseapp.com",
    "databaseURL": "https://telecom-tower-performance-1-default-rtdb.firebaseio.com",
    "projectId": "telecom-tower-performance-1",
    "storageBucket": "telecom-tower-performance-1.appspot.com",
    "messagingSenderId": "681536968586",
    "appId": "1:681536968586:web:82ebaf8bbc7a17e0191a73",
    "measurementId": "G-VP8S76MR98"
}

firebase = pyrebase.initialize_app(firebaseConfig)

# Initialize Firebase Admin SDK with credentials
cred = credentials.Certificate("telecom-tower-performance-1-firebase-adminsdk-76b3k-265f93b36b.json")
firebase_admin.initialize_app(cred, {
    "apiKey": "AIzaSyBiZWWI4vdWhNjwK1547w6dRHKBVZ9it24",
    "authDomain": "telecom-tower-performance-1.firebaseapp.com",
    "databaseURL": "https://telecom-tower-performance-1-default-rtdb.firebaseio.com",
    "projectId": "telecom-tower-performance-1",
    "storageBucket": "telecom-tower-performance-1.appspot.com",
    "messagingSenderId": "681536968586",
    "appId": "1:681536968586:web:82ebaf8bbc7a17e0191a73",
    "measurementId": "G-VP8S76MR98"
})

os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = "glassy-vial-413609-104ff6f112ad.json"

# Initialize Firestore client
db = firestore.client()
auther = firebase.auth()



userDetailsRef = db.collection("UserDetails")
projectRef = db.collection('Projects')

# Password regex pattern
password_regex = re.compile(r'^(?=.*[A-Z])(?=.*\d{2})(?=.*[!@#$%^&()-+=])[A-Za-z\d!@#$%^&*()-+=]{6,}$')
app.config['MAX_CONTENT_LENGTH'] = 30 * 1024 * 1024

if os.path.exists('uploads'):
    shutil.rmtree('uploads')

if not os.path.exists('uploads'):
    os.makedirs('uploads')

UPLOADS_DIR = 'uploads/Predata_RAR'
if not os.path.exists(UPLOADS_DIR):
    os.makedirs(UPLOADS_DIR)

wb = openpyxl.Workbook()
ws = wb.active
excel_file_path = os.path.join('uploads', 'images.xlsx')
wb.save(excel_file_path)
wb = openpyxl.load_workbook(excel_file_path)

POSTUPLOADS_DIR = "postuploads/Postdata_RAR"
post_excel_file_path = "postuploads/postimages.xlsx"

if not os.path.exists('postuploads'):
    os.makedirs('postuploads')

if not os.path.exists(POSTUPLOADS_DIR):
    os.makedirs(POSTUPLOADS_DIR)

wb = openpyxl.Workbook()
ws = wb.active
wb.save(post_excel_file_path)
wb = openpyxl.load_workbook(post_excel_file_path)


@app.route("/")
def home():
    return redirect(url_for('welcome'))

@app.route("/welcome")
def welcome():
    return render_template('welcome.html')

@app.route("/register", methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        employee_name = request.form['employee_name']
        circle_name = request.form['circle_name']
        email = request.form['email']
        password = request.form['password']
        confirm_password = request.form['confirmPassword']

        # Check if passwords match
        if password != confirm_password:
            flash("Passwords do not match. Please try again.", "error")
            return render_template('signup.html')

        # Validate password format
        if not password_regex.match(password):
            flash(
                "Password should contain at least six characters, one uppercase letter, two digits, and one special symbol.",
                "error")
            return render_template('signup.html')

        try:
            user = auther.create_user_with_email_and_password(email, password)
            userDetailsRef.document(user['localId']).set({
                'employee_Name': employee_name,
                'circle_name': circle_name,
                'email': email,
                'isAdmin': False
            })
            flash("Registration successful! You can now login.", "success")
            return redirect(url_for('login'))  # Redirect to login page after successful registration
        except Exception as e:
            error_message = str(e)
            if 'EMAIL_EXISTS' in error_message:
                flash("Email already exists. Please choose a different one.", "error")
                return render_template('signup.html')
            else:
                print('Error creating user:', e)
                flash("Registration failed. Please try again.", "error")
                return render_template('signup.html')
    return render_template('signup.html')

@app.route("/login", methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']

        try:
            user = auther.sign_in_with_email_and_password(email, password)
            session['uid'] = user['localId']
            user_data = db.collection('UserDetails').document(user['localId']).get().to_dict()
            if user_data.get('isAdmin', True):
                return redirect(url_for('welcome_admin'))  # Redirect to welcome_admin page for admin users
            else:
                return redirect(url_for('welcome_user'))  # Redirect to welcome_user page for regular users

        except auth.UserNotFoundError:
            flash("User not found.", "error")
        except ConnectionError:
            flash(" Please enter correct Details", "error")
        except Exception as e:
            print('Error logging in:', e)
            flash("Please enter correct Details", "error")

    return render_template('login.html')
@app.route("/welcomeadmin")
def welcome_admin():
    return render_template('welcomeadmin.html')




@app.route("/welcomeuser")
def welcome_user():
    return render_template('welcomeuser.html')

@app.route("/welcome")
def show_welcome():
    return render_template('welcome.html')

@app.route("/taskallocation", methods=['GET', 'POST'])
def task_allocation():
    global globalCircleName
    if request.method == (''
                          'POST'):
        circleName = request.form.get('project')
        globalCircleName = circleName
        return redirect(url_for('task_allocation'))
    # Fetch employee names from Firestore
    employee_names = []
    users_ref = db.collection('UserDetails').where("circle_name", "==", globalCircleName)
    docs = users_ref.stream()
    for doc in docs:
        employee_names.append(doc.to_dict().get('employee_Name'))
    # Pass employee names to the taskallocation.html template
    return render_template('Taskallocation.html', employee_names=employee_names)

@app.route('/projectallocation')
def projectallocation():
    if request.method == 'POST':
        return redirect(url_for('project_allocation'))
    return render_template('projectallocation.html')

@app.route('/Allocationrequest.html')
def Allocationrequest():
    return render_template('Allocationrequest.html')


@app.route('/AllocatedStatus.html', methods=['GET', 'POST'])
def allocated_status():
    global site_code
    if 'uid' in session:
        uid = session['uid']
        user_ref = db.collection('UserDetails').document(uid)
        user = user_ref.get()

        if user.exists:
            employee_name = user.get("employee_Name")
            print(employee_name)
            doc_snap = db.collection('Projects').where("AssignedEmployee", "==", employee_name).stream()
            print(doc_snap)

            # Iterate over the generator to retrieve documents
            site_codes_list = []
            for doc in doc_snap:
                print("in for loop")
                print(doc.to_dict())
                site_codes_list.append(doc.get("siteCode"))

            print(site_codes_list)

            if request.method == 'POST':
                site_id = request.form.get('siteId')
                site_code=site_id
                print("Site ID:", site_id)  # Debug print

                data={}

                # Fetch the document directly using the site_id
                doc_ref = db.collection('Projects').document(site_id).get()
                if doc_ref.exists:
                    data = doc_ref.to_dict()
                    print("Data fetched:", data)  # Debug print
                    return jsonify(data)
                else:
                    print("No matching document found for site ID:", site_id)
                    return jsonify({})  # Return an empty JSON object if no matching document is found

            return render_template('AllocatedStatus.html', site_codes=site_codes_list)

    # Return a JSON response with an error message if the user is not authenticated
    return jsonify({'error': 'User not authenticated'}), 401

@app.route("/handle_selection", methods=['POST'])
def handle_selection():
    global employeeName
    if request.method == 'POST':
        data = request.get_json()
        selection = data.get('selection')

        # Handle the selected value here


        # Redirect to AllocatedStatus route with selection value
        return redirect(url_for('AllocatedStatus', selection=selection))
    else:
        return jsonify({"error": "Method not allowed"}), 405



@app.route("/submit", methods=['POST'])
def submit():
    if request.method == 'POST':
        project = request.form.get('selectProject')
        operator = request.form.get('selectOperator')
        circle = request.form.get('project')
        activity = request.form.get('selectActivity')
        # Retrieve additional field value if present
        additional_field = request.form.get('hiddenChosenOption', None)

        # Data to be saved in Firestore
        data = {
            "project": project,
            "operator": operator,
            "circle": circle,
            "activity": activity,
            "additional_field": additional_field
        }

        try:
            # Create a new document with a unique ID in the "Project_Selection" collection
            doc_ref = db.collection("Project_Selection").document()
            doc_ref.set(data)
            doc_id = doc_ref.id
            # flash("Form submitted successfully!", "success")
            return redirect(url_for('get_employees', project=project, nameCircle=circle))
        except Exception as e:
            print("Error:", str(e))
            flash("An error occurred in Firestore.", "error")
            return redirect(url_for('projectallocation'))  # Redirect back to the form page in case of an error

    # Handle the case when the form is not submitted via POST method
    return redirect(url_for('projectallocation'))

@app.route("/getEmployees", methods=['POST', 'GET'])
def get_employees():
    circleName = request.args.get('nameCircle')
    project = request.args.get('project')


    # Fetch employees based on the selected circle from Firestore
    employee_names = []
    users_ref = db.collection('UserDetails').where('circle_name', '==', circleName).stream()
    for doc in users_ref:
        employee_names.append(doc.to_dict()['employee_Name'])

    return render_template("Taskallocation.html", project=project, circleName=circleName, employee_names=employee_names)

@app.route("/saveTableData", methods=['POST'])
def save_table_data():
    global globalCircleName
    if request.method == 'POST':
        data = request.json
        print(data)
        try:
            # Save the table data to Firestore
            doc_ref = db.collection("Projects").document(data['siteCode'])
            doc_ref.set(data)
            #flash("Table data saved successfully!", "success")

            # Redirect to fillpostdata.html
            return redirect(url_for('fillpostdata'))
        except Exception as e:
            print("Error:", str(e))
            flash("An error occurred while saving table data.", "error")

        return redirect(url_for('fillpostdata'))



@app.route("/admintaskStatus", methods=["GET", "POST"])
def task_status():
    if request.method == "POST":
        circle_name = request.form.get('circle')
        print(circle_name)
        from_date = request.form.get('fromDate')
        to_date = request.form.get('toDate')

        try:
            from_date_obj = parser.parse(from_date)
            from_date = from_date_obj.strftime('%d-%m-%Y')
            to_date_obj = parser.parse(to_date)
            to_date = to_date_obj.strftime('%d-%m-%Y')
            dataList = []
            datas = db.collection('Projects') \
                .where('CircleName', '==', circle_name) \
                .where('AllocatedDate', '>=', from_date) \
                .where('AllocatedDate', '<=', to_date) \
                .stream()
            for data in datas:
                dataList.append(data.to_dict())
            print("Datalist from admin is", dataList)

            return render_template('AdminTaskStatus.html', circle_name=circle_name, from_date=from_date, data=dataList)


        except Exception as e:

            # Handle any errors that occur during data retrieval

            error_message = f"Error fetching data from Firestore. Please try again later. {e}"

            if "requires an index" in str(e):
                error_message += " The query requires an index. You can create it here: "

                error_message += "https://console.firebase.google.com/v1/r/project/telecom-tower-performance-1/firestore/indexes?create_composite=Clxwcm9qZWN0cy90ZWxlY29tLXRvd2VyLXBlcmZvcm1hbmNlLTEvZGF0YWJhc2VzLyhkZWZhdWx0KS9jb2xsZWN0aW9uR3JvdXBzL1Byb2plY3RzL2luZGV4ZXMvXxABGg4KCkNpcmNsZU5hbWUQARoKCgZTdGF0dXMQARoRCg1BbGxvY2F0ZWREYXRlEAEaDAoIX19uYW1lX18QAQ"

            return error_message

    # Render the form template for GET requests
    return render_template('AdminTaskStatus.html')

@app.route("/usertaskStatus", methods=["GET", "POST"])
def task_status1():
    uid = session['uid']
    print("UID",uid)
    userData = db.collection("UserDetails").document(uid).get().to_dict()
    print("userData Dictionary",userData)
    employeeName = userData['employee_Name']
    print("employee name", employeeName)
    if request.method == "POST":
        circle_status = request.form.get('circle1')
        circle_name = request.form.get('circle')
        from_date = request.form.get('fromDate')
        to_date = request.form.get('toDate')

        try:
            from_date_obj = parser.parse(from_date)
            from_date = from_date_obj.strftime('%d-%m-%Y')
            to_date_obj = parser.parse(to_date)
            to_date = to_date_obj.strftime('%d-%m-%Y')
            dataList = []
            datas = db.collection('Projects') \
                .where('AssignedEmployee', "==", employeeName) \
                .where('CircleName', '==', circle_name) \
                .where('status', '==', circle_status) \
                .where('AllocatedDate', '>=', from_date) \
                .where('AllocatedDate', '<=', to_date) \
                .stream()
            for data in datas:
                dataList.append(data.to_dict())
            print("Datalist from user is", dataList)

            return render_template('UserTaskStatus.html', circle_name=circle_name, from_date=from_date, to_date=to_date, data=dataList)


        except Exception as e:

            error_message = f"Error fetching data from Firestore. Please try again later. {e}"

            if "requires an index" in str(e):
                error_message += " The query requires an index. You can create it here: "

                error_message += "https://console.firebase.google.com/v1/r/project/telecom-tower-performance-1/firestore/indexes?create_composite=Clxwcm9qZWN0cy90ZWxlY29tLXRvd2VyLXBlcmZvcm1hbmNlLTEvZGF0YWJhc2VzLyhkZWZhdWx0KS9jb2xsZWN0aW9uR3JvdXBzL1Byb2plY3RzL2luZGV4ZXMvXxABGg4KCkNpcmNsZU5hbWUQARoKCgZTdGF0dXMQARoRCg1BbGxvY2F0ZWREYXRlEAEaDAoIX19uYW1lX18QAQ"

            return error_message

    # Render the form template for GET requests
    return render_template('UserTaskStatus.html')



@app.route("/nooption.html")
def no_option():
    return render_template("nooption.html")

@app.route("/update_availability", methods=['POST'])
def update_availability():
    if 'uid' not in session:
        return jsonify({'error': 'User not authenticated'}), 401  # Unauthorized

    try:
        data = request.json
        response = data.get('response')
        uid = session['uid']

        if response == 'yes':
            db.collection('UserDetails').document(uid).update({'isAvailable': True, 'Issue': ""})
            return jsonify({'message': 'Availability updated successfully'}), 200
        else:
            db.collection('UserDetails').document(uid).update({'isAvailable': False})
            return jsonify({'message':  'Availability updated successfully'}), 200
    except Exception as e:
        print(e)
        return jsonify({'error': 'Failed to update availability', 'details': str(e)}), 500


@app.route("/nooption.html", methods=['GET', 'POST'])
def store_issue():
    print("Enter in store issues")
    if 'uid' not in session:
        return redirect(url_for('login'))

    if request.method == 'POST':
        issue_text = request.form['other']
        try:
            uid = session.get('uid')  # Get user ID from session
            if uid:
                user_ref = db.collection('UserDetails').document(uid)
                user_ref.set({'Issue': issue_text}, merge=True)
                return redirect(url_for('welcome_user'))
            else:
                return jsonify({'error': 'User not authenticated'}), 401  # Unauthorized
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    else:
        return render_template("nooption.html")



@app.route("/Completestatus")
def Complete_status():
    if request.method == 'POST':
        return redirect(url_for('Complete_status'))
    return render_template('Completestatus.html')

@app.route("/Pendingstatus")
def Pending_status():
    if request.method == 'POST':
        return redirect(url_for('Pending_status'))
    return render_template('pendingstatus.html')

@app.route("/Userdetails")
def User_details():
    if 'uid' not in session:
        return redirect(url_for('login'))

    try:
        uid = session['uid']
        user_data = db.collection('UserDetails').document(uid).get().to_dict()
        if user_data:
            employee_name = user_data.get('employee_Name', '')
            circle_name = user_data.get('circle_name', '')
            email = user_data.get('email', '')
            is_admin = user_data.get('isAdmin')
            if is_admin:
                role = "Admin"
                task_status_link = "/admintaskStatus"  # Admin Task Status Link
            else:
                role = "User"
                task_status_link = "/usertaskStatus"  # User Task Status Link
            return render_template('Userdetails.html', employee_name=employee_name, circle_name=circle_name, email=email, is_admin=is_admin, role=role, task_status_link=task_status_link)
        else:
            return redirect(url_for('login'))
    except Exception as e:
        print('Error fetching user details:', e)
        flash("An error occurred while fetching user details.", "error")
        return redirect(url_for('login'))


@app.route("/request_reset_password", methods=['GET', 'POST'])
def request_reset_password():
    if request.method == 'POST':
        email = request.form.get('email')
        try:
            auther.send_password_reset_email(email)
            flash("Password resent link sent to your email", "success")
            return redirect(url_for('login'))
        except Exception as e:
            return redirect(url_for('request_reset_password'))

    return render_template('ChangePass.html')


@app.route("/DownloadReport")
def report():
    return render_template("DownloadReport.html")


@app.route("/downloadreport", methods=['GET', 'POST'])
def download_report():
    if request.method == 'POST':
        print("Inside If Block")
        site_id = request.form.get('siteId')
        print("Site ID", site_id)
        selected_file = request.form.get('selectFile')
        print("Selected File", selected_file)
        if not site_id or not selected_file:
            flash("Site ID or Selected File is missing.")
            return render_template('/DownloadReport')

        file_data = None

        if selected_file in ['Post_Excel_File_URL', 'Post_Zip_File_URL']:
            file_data = db.collection('Projects').document(site_id).collection('ParameterData').document(
                'PostData').get()
        elif selected_file in ['Pre_Excel_File_URL', 'Pre_Zip_File_URL']:
            file_data = db.collection('Projects').document(site_id).collection('ParameterData').document(
                'PreData').get()

        if file_data:
            file_url = file_data.get(selected_file)
            if file_url:
                try:
                    return redirect(file_url)
                except Exception as e:
                    print(e)
                    flash("Error while downloading file.")
            else:
                flash("Selected file does not exist.")
        else:
            flash("Site data does not exist.")

    return redirect('/DownloadReport')

@app.route("/userrequests")
def user_requests():
    return render_template('userreques.html')

@app.route('/fillpostdata')
def fillpostdata():
    return render_template('fillpostdata.html')

@app.route('/fillpostdata', methods=['POST'])
def submit_form():
    try:
        # Get form data
        site_id = request.form.get('siteId')
        sector = request.form.get('sector')
        azimuth = request.form.get('azimuth')
        tower_height = request.form.get('towerHeight')
        mechanical_cell = request.form.get('mechanicalCell')
        electrical_cell = request.form.get('electricalCell')
        pole_tilt = request.form.get('poleTilt')
        antennaheight = request.form.get('antennaheight')

        doc_ref = db.collection('Projects').document(site_id).collection("ParameterData").document(
            "PostData").collection(sector).document("Requirement")
        if not doc_ref.get().exists:
            doc_ref.set({})
        doc_ref.set({
            'azimuth': azimuth,
            'tower_height': tower_height,
            'mechanical_cell': mechanical_cell,
            'electrical_cell': electrical_cell,
            'pole_tilt': pole_tilt,
            'antennaheight': antennaheight
        })

        return redirect(url_for('fillpostdata'))
    except Exception as e:
        print("Error submitting form:", e)
        return "An error occurred while submitting the form."


@app.route('/logdetails', methods=['GET', 'POST'])
def logdetails():
    site_code = request.args.get('siteCode')
    print("log detail function started")
    all_data = []
    sectors = ['sector1', 'sector2', 'sector3']
    datas = ['PreData', 'PostData']
    for data_item in datas:
        print("Data_Item",data_item)
        for sector in sectors:
            print("site_id:", site_code)
            print("data:", data_item)
            print("sector:", sector)
            site_doc_ref = db.collection("Projects").document(site_code).collection("ParameterData").document(data_item).collection(sector).document("Status")
            print("Path", site_doc_ref.path)
            log_details_data = site_doc_ref.get().to_dict()
            if log_details_data:
                sector_name = sector
                for key, value in log_details_data.items():
                    if key != 'Technology':
                        status = "Done" if value else "Pending"
                        all_data.append({
                            "SiteID": site_code,
                            "Sector": sector_name,
                            "Field": key,
                            "Type": data_item,
                            "Value": value,
                            "Status": status
                        })
            print("Log data is:", log_details_data)

    # Render template with fetched data
    print("log details exited")
    return render_template('logdetails.html', data=all_data)

@app.route('/presectorselectionpage.html')
def presectorselectionpage():
    print("presectorselectionpage")
    global site_code
    print(site_code)
    return render_template('presectorselectionpage.html')

@app.route('/presector1.html')
def index():
    return render_template('presector1.html')

@app.route('/presector2.html')
def presector2():
    return render_template('presector2.html')

@app.route('/presector3.html')
def presector3():
    return render_template('presector3.html')

@app.route('/upload-images-1', methods=['POST'])
def upload_images_1():
    global site_code
    print(site_code)
    image_keys = ['AzimuthCellSec1', 'MechanicalSec1', 'ElectricalSec1', 'AntennaHeightSec1',
                  'AntBuildingSec1', 'BuildHeightSec1', 'PoleTiltSec1', 'MirrorCompassSec1', 'AntennaMarkingSec1']
    save_images(image_keys, 'sector1', site_code)
    return redirect(url_for("presector2"))


@app.route('/upload-images-2', methods=['POST'])
def upload_images_2():
    global site_code
    print(site_code)
    image_keys = ['AzimuthCellSec2', 'MechanicalSec2', 'ElectricalSec2', 'AntennaHeightSec2',
                  'AntBuildingSec2', 'BuildHeightSec2', 'PoleTiltSec2', 'MirrorCompassSec2', 'AntennaMarkingSec2']
    save_images(image_keys, 'sector2', site_code)
    return redirect(url_for("presector3"))


@app.route('/upload-images-3', methods=['POST'])
def upload_images_3():
    global site_code
    print(site_code)
    image_keys = ['AzimuthCellSec3', 'MechanicalSec3', 'ElectricalSec3', 'AntennaHeightSec3',
                  'AntBuildingSec3', 'BuildHeightSec3', 'PoleTiltSec3', 'MirrorCompassSec3', 'AntennaMarkingSec3']
    save_images(image_keys, 'sector3', site_code)

    # Zip the Predata_RAR folder
    shutil.make_archive(UPLOADS_DIR, 'zip', UPLOADS_DIR)
    print("upload_images3")
    return redirect(url_for("post_data"))
def save_images(image_keys, sec, siteCode):
    images = {}
    counter=1
    cellCount=1

    uploaded_images = {}

    for key in image_keys:
        print(key)
        print(sec)
        file = request.files.get(key)
        if file is not None:  # Check if an image is uploaded
            print("If started")
            images[key] = file
            uploaded_images[key] = True  # Mark the key as True if image is uploaded
        else:
            print("Else started")
            images[key] = None
            uploaded_images[key] = False
    # Your existing code for Excel manipulation...

    db = firestore.client()
    project_ref = db.collection('Projects').document(siteCode)

    parameter_data_ref = project_ref.collection('ParameterData')
    pre_data_ref = parameter_data_ref.document('PreData')
    sec_collection_ref = pre_data_ref.collection(sec)
    status_doc_ref = sec_collection_ref.document('Status')

    # Update the status document with image upload status for each key
    status_doc_ref.set(uploaded_images)

    print("Data updated in Firestore successfully")

    ws = wb.create_sheet(title=sec, index=0)

    for key, file in images.items():
        if file:
            # Save the file in the uploads directory
            file_path = os.path.join(UPLOADS_DIR, f"{sec}_{counter}.jpg")
            file.save(file_path)

            # Add image metadata to the Excel sheet
            cell = ws[f'A{cellCount}']
            cell.value = str(image_keys[counter - 1])
            cell.font = Font(size='16', bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[cellCount].height = 300
            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['A'].width = 20

            # Add image to the Excel sheet
            img = openpyxl.drawing.image.Image(file_path)
            img.width = 250
            img.height = 400
            img.anchor = f'B{cellCount}'
            ws.add_image(img)

        cellCount += 2
        counter += 1

    # Save the modified Excel file
    wb.save(excel_file_path)

    # Upload Excel file to Firebase Storage
    bucket = storage.bucket()
    excel_blob = bucket.blob('pre_data/images.xlsx')
    excel_blob.upload_from_filename(excel_file_path)
    excel_url = excel_blob.public_url

    # Create a zip file of the images directory
    shutil.make_archive(UPLOADS_DIR, 'zip', UPLOADS_DIR)

    # Upload the zip file to Firebase Storage
    zip_blob = bucket.blob('zipF/Predata_RAR.zip')
    zip_blob.upload_from_filename('uploads/Predata_RAR.zip')
    zip_url = zip_blob.public_url

    print("Files uploaded successfully to Firebase Storage")
    pre_save_url_to_firestore(excel_url, zip_url)

def pre_save_url_to_firestore(excel_url, zip_url):
    today_date = datetime.today().strftime('%d-%m-%Y')

    # Create a dictionary with document data
    document_data = {
        "date": today_date,
        "Pre_Excel_File_URL": excel_url,
        "Pre_Zip_File_URL": zip_url,
    }

    db.collection("Projects").document(site_code).collection("ParameterData").document("PreData").set(document_data)

    print("File URLs and date saved to Firestore.")


def upload_to_storage():
    bucket = storage.bucket()

    # Upload Excel file
    excel_blob = bucket.blob('Audit_Data/')
    excel_blob.upload_from_filename(excel_file_path)

    # Upload zipped folder
    zip_blob = bucket.blob('Audit_Data/')
    zip_blob.upload_from_filename('uploads/Predata_RAR.zip')

    print("Files uploaded successfully to Firebase Storage")

    # Generate the document name

    def save_urls_to_firestore(excel_url, zip_url):
        today_date = datetime.today().strftime('%d-%m-%Y')

        # Create a dictionary with document data
        document_data = {
            "date": today_date,
            "Download_Excel_Data": excel_url,
            "Download_Zip_Data": zip_url,
        }

        # Set the document in Firestore with the provided name
        db.collection("files").document(today_date).set(document_data)

        print("File URLs and document name saved to Firestore.")

    def main():
        excel_file_path = 'path/to/excel_file.xlsx'
        document_name = input("Enter the site_id to store the file in: ")

        # Upload files to Firebase Storage
        excel_url, zip_url = upload_to_storage(excel_file_path)

        # Save URLs to Firestore
        save_urls_to_firestore(excel_url, zip_url)

    def upload_to_storage(excel_file_path):
        bucket = storage.bucket()

    # Upload Excel file
    excel_blob_name = 'pre_data/images.xlsx'
    excel_blob = bucket.blob(excel_blob_name)
    excel_blob.upload_from_filename(excel_file_path)

    # Create a zip file of the images directory
    uploads_dir = 'path/to/uploads/directory'  # Modify as needed
    shutil.make_archive(uploads_dir, 'zip', uploads_dir)

    # Upload zipped folder
    zip_blob_name = 'zipF/Predata_RAR.zip'
    zip_blob = bucket.blob(zip_blob_name)
    zip_blob.upload_from_filename(f'{uploads_dir}.zip')

    # Get the URLs of the uploaded files
    excel_url = excel_blob.public_url
    zip_url = zip_blob.public_url

    print("Files uploaded successfully to Firestore")
    return excel_url,zip_url

@app.route('/postdata.html', methods=['GET', 'POST'])
def post_data():
    if 'uid' in session:
        uid = session['uid']
        user_ref = db.collection('UserDetails').document(uid)
        user = user_ref.get()

        if user.exists:
            # Fetch the site IDs assigned to the current user
            employee_name = user.get("employee_Name")
            doc_snap = db.collection('Projects').where("AssignedEmployee", "==", employee_name).stream()
            site_codes_list = [doc.get("siteCode") for doc in doc_snap]

            if request.method == 'POST':
                site_id = request.form.get('siteId')
                sector = request.form.get('sector')

                data = {"SiteID": site_id}
                doc_ref = db.collection('Projects').document(site_id).collection('ParameterData').document('PostData').collection(sector).document('Requirement')
                doc = doc_ref.get()
                if doc.exists:
                    data.update(doc.to_dict())
                    print("Data fetched:", data)  # Debug print

                return jsonify(data)

            return render_template('postdata.html', site_codes=site_codes_list)
    return "Unauthorized", 401  # If user is not logged in or not authorized

@app.route('/postsectorselection.html')
def post_sector_selection():
    site_code = request.args.get('SiteID')
    session['site_code'] = site_code  # Store SiteID in session
    print("postsectorselection siteID", site_code)
    return render_template("postsectorselection.html")


@app.route('/postsector1.html')
def postsector1():
    return render_template("postsector1.html")


@app.route('/postsector2.html')
def postsector2():
    return render_template("postsector2.html")


@app.route('/postsector3.html')
def postsector3():
    return render_template("postsector3.html")


@app.route('/extract_text', methods=['POST'])
def extract_text():
    if 'image' not in request.files:
        return jsonify({'error': 'No image provided'}), 400

    image_file = request.files['image']
    image_content = image_file.read()

    client = vision.ImageAnnotatorClient()
    image = vision.Image(content=image_content)
    response = client.text_detection(image=image)
    texts = response.text_annotations

    if texts:
        extracted_text = texts[0].description
        return jsonify({'text': extracted_text}), 200
    else:
        return jsonify({'text': 'No text found in the image.'}), 404



@app.route('/postupload-images-1', methods=['POST'])
def postupload_images_1():
    site_code = session.get('site_code')  # Retrieve SiteID from session
    print("SiteID:", site_code)  # Print SiteID to console
    if not site_code:
        return jsonify({'error': 'SiteID not found in session'}), 400

    postimage_keys = ['AzimuthCellSec1', 'MechanicalSec1', 'ElectricalSec1', 'AntennaHeightSec1',
                      'AntBuildingSec1', 'BuildHeightSec1', 'TowerHeightSec1', 'PoleTiltSec1', 'MirrorCompassSec1', 'AntennaMarkingSec1']

    # Validate Azimuth image
    azimuth_file = request.files.get('AzimuthCellSec1')
    if azimuth_file:
        print("Enters  in azimuth")
        client = vision.ImageAnnotatorClient()
        image_content = azimuth_file.read()
        image = vision.Image(content=image_content)
        response = client.text_detection(image=image)
        texts = response.text_annotations
        print(texts)

        if texts:
            print("enters in text")
            extracted_text = texts[0].description
            doc_ref = db.collection('Projects').document(site_code).collection('ParameterData').document('PostData').collection('sector1').document('Requirement')
            print(doc_ref)
            doc = doc_ref.get()
            if doc.exists:
                print("enterindocs")
                data = doc.to_dict()
                expected_text = data.get('azimuth', '')
                if expected_text not in extracted_text:
                    return jsonify({'error': 'Invalid Azimuth image'}), 400
        else:
            return jsonify({'error': 'No text found in the Azimuth image.'}), 400

    # Validate AntennaHeight image
    antenna_height_file = request.files.get('AntennaHeightSec1')
    if antenna_height_file:
        print("Enter in antenna height")
        client = vision.ImageAnnotatorClient()
        image_content = antenna_height_file.read()
        image = vision.Image(content=image_content)
        response = client.text_detection(image=image)
        texts = response.text_annotations

        if texts:
            print("Enters in text")
            extracted_text = texts[0].description
            doc_ref = db.collection('Projects').document(site_code).collection('ParameterData').document('PostData').collection('sector1').document('Requirements')
            doc = doc_ref.get()
            if doc.exists:
                print("enter in docs")
                data = doc.to_dict()
                expected_text = data.get('antenna_height', '')
                if expected_text not in extracted_text:
                    return jsonify({'error': 'Invalid AntennaHeight image'}), 400
        else:
            return jsonify({'error': 'No text found in the AntennaHeight image.'}), 400

    build_height_file = request.files.get('BuildHeightSec1')
    if build_height_file:
        client = vision.ImageAnnotatorClient()
        image_content = build_height_file.read()
        image = vision.Image(content=image_content)
        response = client.text_detection(image=image)
        texts = response.text_annotations

        if texts:
            extracted_text = texts[0].description
            doc_ref = db.collection('Projects').document(site_code).collection('ParameterData').document('PostData').collection('sector1').document('Requirements')
            doc = doc_ref.get()
            if doc.exists:
                data = doc.to_dict()
                expected_text = data.get('build_height', '')
                if expected_text not in extracted_text:
                    return jsonify({'error': 'Invalid BuildHeight image'}), 400
        else:
            return jsonify({'error': 'No text found in the BuildHeight image.'}), 400

    # Validate TowerHeight image
    tower_height_file = request.files.get('TowerHeightSec1')
    if tower_height_file:
        client = vision.ImageAnnotatorClient()
        image_content = tower_height_file.read()
        image = vision.Image(content=image_content)
        response = client.text_detection(image=image)
        texts = response.text_annotations

        if texts:
            extracted_text = texts[0].description
            doc_ref = db.collection('Projects').document(site_code).collection('ParameterData').document('PostData').collection('sector2').document('Requirements')
            doc = doc_ref.get()
            if doc.exists:
                data = doc.to_dict()
                expected_text = data.get('tower_height', '')
                if expected_text not in extracted_text:
                    return jsonify({'error': 'Invalid TowerHeight image'}), 400
        else:
            return jsonify({'error': 'No text found in the TowerHeight image.'}), 400

    # If neither validation fails, proceed with saving other images
    postsave_images(postimage_keys, 'sector1', site_code)
    print("saving post image")
    return redirect(url_for("postsector2"))

@app.route('/postupload-images-2', methods=['POST'])
def postupload_images_2():
    site_code = session.get('site_code')  # Retrieve SiteID from session
    print("SiteID:", site_code)  # Print SiteID to console
    if not site_code:
        return jsonify({'error': 'SiteID not found in session'}), 400

    postimage_keys = ['AzimuthCellSec2', 'MechanicalSec2', 'ElectricalSec2', 'AntennaHeightSec2',
                      'AntBuildingSec2', 'BuildHeightSec2', 'PoleTiltSec2', 'MirrorCompassSec2', 'AntennaMarkingSec2']

    # Validate Azimuth image for sector 2
    azimuth_file = request.files.get('AzimuthCellSec2')
    if azimuth_file:
        print("Enters  in azimuth sector2")
        client = vision.ImageAnnotatorClient()
        image_content = azimuth_file.read()
        image = vision.Image(content=image_content)
        response = client.text_detection(image=image)
        texts = response.text_annotations
        print(texts)

        if texts:
            print("enters in text sector2")
            extracted_text = texts[0].description
            doc_ref = db.collection('Projects').document(site_code).collection('ParameterData').document(
                'PostData').collection('sector2').document('Requirement')
            print(doc_ref)
            doc = doc_ref.get()
            if doc.exists:
                print("enterindocs sector2")
                data = doc.to_dict()
                expected_text = data.get('azimuth', '')
                if expected_text not in extracted_text:
                    return jsonify({'error': 'Invalid Azimuth image'}), 400
        else:
            return jsonify({'error': 'No text found in the Azimuth image.'}), 400

    # Validate AntennaHeight image for sector 2
    antenna_height_file = request.files.get('AntennaHeightSec2')
    if antenna_height_file:
        client = vision.ImageAnnotatorClient()
        image_content = antenna_height_file.read()
        image = vision.Image(content=image_content)
        response = client.text_detection(image=image)
        texts = response.text_annotations

        if texts:
            extracted_text = texts[0].description
            doc_ref = db.collection('Projects').document(site_code).collection('ParameterData').document('PostData').collection('sector2').document('Requirements')
            doc = doc_ref.get()
            if doc.exists:
                data = doc.to_dict()
                expected_text = data.get('antenna_height', '')
                if expected_text not in extracted_text:
                    return jsonify({'error': 'Invalid AntennaHeight image for sector 2'}), 400
        else:
            return jsonify({'error': 'No text found in the AntennaHeight image for sector 2.'}), 400

    build_height_file = request.files.get('BuildHeightSec2')
    if build_height_file:
        client = vision.ImageAnnotatorClient()
        image_content = build_height_file.read()
        image = vision.Image(content=image_content)
        response = client.text_detection(image=image)
        texts = response.text_annotations

        if texts:
            extracted_text = texts[0].description
            doc_ref = db.collection('Projects').document(site_code).collection('ParameterData').document('PostData').collection('sector2').document('Requirements')
            doc = doc_ref.get()
            if doc.exists:
                data = doc.to_dict()
                expected_text = data.get('build_height', 'InstaICT')
                if expected_text not in extracted_text:
                    return jsonify({'error': 'Invalid BuildHeight image for sector 2'}), 400
        else:
            return jsonify({'error': 'No text found in the BuildHeight image for sector 2.'}), 400

    postsave_images(postimage_keys, 'sector2', site_code)
    return redirect(url_for("postsector3"))


@app.route('/postupload-images-3', methods=['POST'])
def postupload_images_3():
    site_code = session.get('site_code')  # Retrieve SiteID from session
    print("SiteID:", site_code)  # Print SiteID to console
    if not site_code:
        return jsonify({'error': 'SiteID not found in session'}), 400
    postimage_keys = ['AzimuthCellSec3', 'MechanicalSec3', 'ElectricalSec3', 'AntennaHeightSec3',
                      'AntBuildingSec3', 'BuildHeightSec3', 'PoleTiltSec3', 'MirrorCompassSec3', 'AntennaMarkingSec3']

    # Validate Azimuth image for sector 3
    azimuth_file = request.files.get('AzimuthCellSec3')
    if azimuth_file:
        print("Enters  in azimuth sector3")
        client = vision.ImageAnnotatorClient()
        image_content = azimuth_file.read()
        image = vision.Image(content=image_content)
        response = client.text_detection(image=image)
        texts = response.text_annotations
        print(texts)

        if texts:
            print("enters in text")
            extracted_text = texts[0].description
            doc_ref = db.collection('Projects').document(site_code).collection('ParameterData').document(
                'PostData').collection('sector3').document('Requirement')
            print(doc_ref)
            doc = doc_ref.get()
            if doc.exists:
                print("enterindocs sector3")
                data = doc.to_dict()
                expected_text = data.get('azimuth', '')
                if expected_text not in extracted_text:
                    return jsonify({'error': 'Invalid Azimuth image'}), 400
        else:
            return jsonify({'error': 'No text found in the Azimuth image.'}), 400

    # Validate AntennaHeight image for sector 3
    antenna_height_file = request.files.get('AntennaHeightSec3')
    if antenna_height_file:
        client = vision.ImageAnnotatorClient()
        image_content = antenna_height_file.read()
        image = vision.Image(content=image_content)
        response = client.text_detection(image=image)
        texts = response.text_annotations

        if texts:
            extracted_text = texts[0].description
            doc_ref = db.collection('Projects').document(site_code).collection('ParameterData').document('PostData').collection('sector3').document('Requirements')
            doc = doc_ref.get()
            if doc.exists:
                data = doc.to_dict()
                expected_text = data.get('antenna_height', '')
                if expected_text not in extracted_text:
                    return jsonify({'error': 'Invalid AntennaHeight image for sector 3'}), 400
        else:
            return jsonify({'error': 'No text found in the AntennaHeight image for sector 3.'}), 400

    build_height_file = request.files.get('BuildHeightSec3')
    if build_height_file:
        client = vision.ImageAnnotatorClient()
        image_content = build_height_file.read()
        image = vision.Image(content=image_content)
        response = client.text_detection(image=image)
        texts = response.text_annotations

        if texts:
            extracted_text = texts[0].description
            doc_ref = db.collection('Projects').document(site_code).collection('ParameterData').document('PostData').collection('sector3').document('Requirements')
            doc = doc_ref.get()
            if doc.exists:
                data = doc.to_dict()
                expected_text = data.get('build_height', '')
                if expected_text not in extracted_text:
                    return jsonify({'error': 'Invalid BuildHeight image for sector 3'}), 400
        else:
            return jsonify({'error': 'No text found in the BuildHeight image for sector 3.'}), 400
    postsave_images(postimage_keys, 'sector3', site_code)

    # Zip the Predata_RAR folder
    shutil.make_archive(POSTUPLOADS_DIR, 'zip', POSTUPLOADS_DIR)
    return redirect(url_for("post_data"))

def postsave_images(postimage_keys, postsec, siteCode):
    postimages = {}
    postcounter = 1
    postcellCount = 1
    site_code = session.get('site_code')  # Retrieve SiteID from session
    print("SiteID:", site_code)  # Print SiteID to console
    if not site_code:
        return jsonify({'error': 'SiteID not found in session'}), 400

    for key in postimage_keys:
        if key in request.files:
            file = request.files[key]
            file_content = file.read()  # Read file content
            img = PILImage.open(file)
            img = img.convert('RGB')  # Convert image to RGB mode
            post_file_path = os.path.join(POSTUPLOADS_DIR, f"{postsec}_{key}.jpg")
            img.save(post_file_path)
            postimages[key] = True  # Mark the key as True if image is uploaded
        else:
            postimages[key] = False

    # Update Firestore with upload status
    db = firestore.client()
    project_ref = db.collection('Projects').document(siteCode)
    parameter_data_ref = project_ref.collection('ParameterData')
    post_data_ref = parameter_data_ref.document('PostData')
    sec_collection_ref = post_data_ref.collection(postsec)
    status_doc_ref = sec_collection_ref.document('Status')
    status_doc_ref.set(postimages)

    print("Data updated in Firestore successfully")
    wb = openpyxl.load_workbook(post_excel_file_path)

    for postkey in postimage_keys:
        file = request.files.get(postkey)
        if file:
            postimages[postkey] = file
        else:
            postimages[postkey] = None

    # Create a new worksheet in the Excel file
    ws = wb.create_sheet(title=postsec, index=0)

    for key, file in postimages.items():
        if file:
            post_file_path = os.path.join(POSTUPLOADS_DIR, f"{postsec}_{postcounter}.jpg")

            try:
                img = PILImage.open(file)
            except PILImage.UnidentifiedImageError:
                print(f"Unsupported image format for {key}: {file.filename}")
                continue

            img.save(post_file_path)

            cell = ws[f'A{postcellCount}']
            cell.value = str(postimage_keys[postcounter - 1])
            cell.font = Font(size='16', bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[postcellCount].height = 300
            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['A'].width = 20

            img = openpyxl.drawing.image.Image(post_file_path)
            img.width = 250
            img.height = 400
            img.anchor = f'B{postcellCount}'
            ws.add_image(img)

            postcellCount += 2
            postcounter += 1

    wb.save(post_excel_file_path)

    bucket = storage.bucket()
    excel_blob = bucket.blob('post_data/postimages.xlsx')
    excel_blob.upload_from_filename(post_excel_file_path)
    excel_url = excel_blob.public_url

    shutil.make_archive(POSTUPLOADS_DIR, 'zip', POSTUPLOADS_DIR)

    zip_blob = bucket.blob('zipF/Postdata_RAR.zip')
    zip_blob.upload_from_filename('postuploads/Postdata_RAR.zip')
    zip_url = zip_blob.public_url

    print("Files uploaded successfully to Firebase Storage")
    post_save_url_to_firestore(excel_url, zip_url)

def post_save_url_to_firestore(excel_url, zip_url):
    site_code = session.get('site_code')  # Retrieve SiteID from session
    print("SiteID:", site_code)  # Print SiteID to console
    if not site_code:
        return jsonify({'error': 'SiteID not found in session'}), 400
    today_date = datetime.today().strftime('%d-%m-%Y')

    # Create a dictionary with document data
    document_data = {
        "date": today_date,
        "Post_Excel_File_URL": excel_url,
        "Post_Zip_File_URL": zip_url,
    }

    db.collection("Projects").document(site_code).collection("ParameterData").document("PostData").set(document_data)

    print("File URLs and date saved to Firestore.")


def upload_to_storage():
    site_code = session.get('site_code')  # Retrieve SiteID from session
    print("SiteID:", site_code)  # Print SiteID to console
    if not site_code:
        return jsonify({'error': 'SiteID not found in session'}), 400
    bucket = storage.bucket()

    # Upload Excel file
    excel_blob = bucket.blob('Audit_Data/')
    excel_blob.upload_from_filename(excel_file_path)

    # Upload zipped folder
    zip_blob = bucket.blob('Audit_Data/')
    zip_blob.upload_from_filename('postuploads/Postdata_RAR.zip')

    print("Files uploaded successfully to Firebase Storage")

    # Generate the document name

    def save_urls_to_firestore(excel_url, zip_url):
        today_date = datetime.today().strftime('%d-%m-%Y')

        # Create a dictionary with document data
        document_data = {
            "date": today_date,
            "Download_Excel_Data": excel_url,
            "Download_Zip_Data": zip_url,
        }

        # Set the document in Firestore with the provided name
        db.collection("Projects").document(site_code).collection("ParameterData").document("PostData").set(document_data)
        projectCompletionData = {
            "status" : "Completed"
        }
        try:
            db.collection("Projects").document(site_code).set(projectCompletionData)
        except Exception as f:
            print("Error", f)
        print("File URLs and document name saved to Firestore.")

    def main():
        excel_file_path = 'path/to/excel_file.xlsx'
        document_name = input("Enter the site_id to store the file in: ")
        # Upload files to Firebase Storage
        excel_url, zip_url = upload_to_storage(excel_file_path)

        # Save URLs to Firestore
        save_urls_to_firestore(excel_url, zip_url)

    def upload_to_storage(excel_file_path):
        bucket = storage.bucket()

    # Upload Excel file
    excel_blob_name = 'post_data/postimages.xlsx'
    excel_blob = bucket.blob(excel_blob_name)
    excel_blob.upload_from_filename(post_excel_file_path)

    # Create a zip file of the images directory
    uploads_dir = 'path/to/uploads/directory'  # Modify as needed
    shutil.make_archive(POSTUPLOADS_DIR, 'zip', POSTUPLOADS_DIR)

    # Upload zipped folder
    zip_blob_name = 'zipF/Postdata_RAR.zip'
    zip_blob = bucket.blob(zip_blob_name)
    zip_blob.upload_from_filename(f'{POSTUPLOADS_DIR}.zip')

    # Get the URLs of the uploaded files
    excel_url = excel_blob.public_url
    zip_url = zip_blob.public_url

    print("Files uploaded successfully to Firestore")
    return excel_url, zip_url



if __name__ == '__main__':
    app.run(debug=True)



